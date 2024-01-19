import openpyxl
from PIL import Image, ImageDraw, ImageFont

# abrir planilha = openpyxl.load_workbook("caminho")
# planilha[aba/planilha].iter_rows(min_row="valor de onde começa") / planilha[aba/planilha].iter_cols(min_col="valor de onde começa")

workbook_students = openpyxl.load_workbook("src/assets/planilha_alunos.xlsx")

for index, row in enumerate(workbook_students["Sheet1"].iter_rows(min_row=2)):
	course_name = row[0].value
	student_name = row[1].value
	participation_type = row[2].value
	start_date = row[3].value
	end_date = row[4].value
	workload = row[5].value
	date_certificate = row[6].value

	name_font = ImageFont.truetype("src/assets/FiraCode-Bold.ttf", 80)
	general_font = ImageFont.truetype("src/assets/FiraCode-Regular.ttf", 70)
	date_font = ImageFont.truetype("src/assets/FiraCode-Regular.ttf", 55)

	certificate_img = Image.open("src/assets/certificado_padrao.jpg")
	draw = ImageDraw.Draw(certificate_img)

	draw.text((1015, 840), student_name, fill="black", font=name_font)
	draw.text((1077, 967), course_name, fill="black", font=general_font)
	draw.text((1435, 1080), participation_type, fill="black", font=general_font)
	draw.text((1495, 1200), str(workload), fill="black", font=general_font)
	draw.text((720, 1785), str(start_date), fill="red", font=date_font)
	draw.text((720, 1935), str(end_date), fill="red", font=date_font)
	draw.text((2195, 1935), str(date_certificate), fill="red", font=date_font)


	certificate_img.save(f"src/img_tests/{index} - {student_name}.png")

